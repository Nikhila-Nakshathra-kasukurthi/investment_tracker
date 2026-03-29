from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
import time

# -------------------------------
# OPEN BROWSER
# -------------------------------
driver = webdriver.Chrome()
driver.get("https://www.chittorgarh.com/report/ipo-in-india-list-main-board-sme/82/all/?year=2026")
driver.maximize_window()
time.sleep(2)

wait = WebDriverWait(driver, 10)

# -------------------------------
# YEAR SELECTION
# -------------------------------
year = "2009"
try:
    year_dropdown = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "cg-select")))
    Select(year_dropdown).select_by_value(year)
    time.sleep(3)
    print(f"Year selected: {year}")
except:
    print("Year dropdown not found")

# -------------------------------
# CLOSE POPUP IF ANY
# -------------------------------
try:
    overlay = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.position-fixed")))
    driver.execute_script("arguments[0].remove();", overlay)
    print("Popup removed successfully.")
    time.sleep(2)
except:
    print("No popup found.")

# -------------------------------
# SET TABLE LENGTH TO 75
# -------------------------------
try:
    length_dropdown = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "select[name^='report_table_length']")))
    Select(length_dropdown).select_by_value("75")
    time.sleep(2)
    print("Table length set to 75 rows per page.")
except:
    print("Table length dropdown not found.")

# -------------------------------
# EXCEL SETUP
# -------------------------------
workbook = Workbook()
sheet = workbook.active
sheet.title = f"IPO_{year}"
all_rows = []

# -------------------------------
# CATEGORY XPATHS
# -------------------------------
category_xpaths = [
    '//*[@id="table_section"]/div[1]/div[1]/div/div/div[1]/li/a',  # All IPO
    '//*[@id="table_section"]/div[1]/div[1]/div/div/div[2]/li/a',  # Mainboard IPO
    '//*[@id="table_section"]/div[1]/div[1]/div/div/div[3]/li/a',  # SME IPO
    '//*[@id="table_section"]/div[1]/div[1]/div/div/div[4]/li/a',  # ReITs IPO
    '//*[@id="table_section"]/div[1]/div[1]/div/div/div[5]/li/a',  # InvITs IPO
    '//a[contains(text(),"Mainboard FPO")]'                        # Mainboard FPO
]

# -------------------------------
# SAFE FIND FUNCTION
# -------------------------------
def safe_find(xpath):
    try:
        return driver.find_element(By.XPATH, xpath)
    except:
        return None

# -------------------------------
# LOOP THROUGH CATEGORIES
# -------------------------------
for cat_xpath in category_xpaths:
    try:
        cat_element = wait.until(EC.element_to_be_clickable((By.XPATH, cat_xpath)))
        cat_name = cat_element.text
        driver.execute_script("arguments[0].click();", cat_element)
        WebDriverWait(driver, 10).until(
            lambda d: len(d.find_elements(By.XPATH, '//*[@id="report_table"]/tbody/tr')) > 0
        )
        time.sleep(1)
    except:
        print(f"Category not found: {cat_xpath}")
        continue

    # -------------------------------
    # INITIALIZE HEADER VARIABLES
    # -------------------------------
    COMPANY = Opening_date = Closing_date = List_path = Issue_price = Total_issue_amount = Listing_At = Lead_manager = Compare = None

    # -------------------------------
    # SCROLL TABLE HEADERS INTO VIEW
    # -------------------------------
    try:
        COMPANY = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@id="report_table"]/thead/tr[1]/th[1]')))
        Opening_date = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@id="report_table"]/thead/tr[1]/th[2]/span[1]')))
        Closing_date = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@id="report_table"]/thead/tr[1]/th[3]/span[1]')))
        List_path = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@id="report_table"]/thead/tr[1]/th[4]/span[1]')))
        Issue_price = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@id="report_table"]/thead/tr[1]/th[5]/span[1]')))
        Total_issue_amount = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@id="report_table"]/thead/tr[1]/th[6]/span[1]')))
        Listing_At = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@id="report_table"]/thead/tr[1]/th[7]/span[1]')))
        Lead_manager = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@id="report_table"]/thead/tr[1]/th[8]/span[1]')))
        Compare = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@id="report_table"]/thead/tr[1]/th[9]/span[1]')))
    except:
        print("Some headers could not be found or scrolled.")

    # -------------------------------
    # PAGE LOOP WITH DUPLICATE CHECK
    # -------------------------------
    page = 1
    seen_rows = set()  # Track unique rows by text
    while True:
        try:
            rows = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="report_table"]/tbody/tr')))
        except:
            print(f"No rows found on page {page} for category {cat_name}")
            break

        # Add headers once
        if len(all_rows) == 0:
            headers = [
                COMPANY, Opening_date, Closing_date, List_path,
                Issue_price, Total_issue_amount, Listing_At,
                Lead_manager, Compare
            ]
            sheet.append(["Year"] + [h.text if h else "N/A" for h in headers])

        # Append row data without duplicates
        for row in rows:
            row_data = [col.text for col in row.find_elements(By.TAG_NAME, "td")]
            row_key = tuple(row_data)  # Use tuple as unique key
            if row_key not in seen_rows:
                seen_rows.add(row_key)
                all_rows.append([year] + row_data)
                sheet.append([year] + row_data)

        print(f"Category '{cat_name}' page {page} Rows added: {len(rows)}")
        page += 1

        # CLICK NEXT BUTTON IF EXISTS
        try:
            next_btn = driver.find_element(By.XPATH, '//ul[@class="pagination"]/li/a[text()="Next"]')
            if "disabled" in next_btn.get_attribute("class"):
                break
            driver.execute_script("arguments[0].click();", next_btn)
            WebDriverWait(driver, 10).until(
                lambda d: len(d.find_elements(By.XPATH, '//*[@id="report_table"]/tbody/tr')) != len(rows)
            )
            time.sleep(1)
        except:
            break

# -------------------------------
# SAVE EXCEL
# -------------------------------
file_name = f"ipo_full_dataset_{int(time.time())}.xlsx"
workbook.save(file_name)
print(f"Excel saved as {file_name}")

# -------------------------------
# FORMAT EXCEL
# -------------------------------
wb = load_workbook(file_name)
ws = wb.active
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(file_name)
print("Excel formatted properly. Total rows scraped:", len(all_rows))

driver.quit()